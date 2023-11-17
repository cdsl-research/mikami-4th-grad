import os
import openpyxl
from openpyxl.styles import PatternFill
import subprocess
import socket
import threading

def get_process_id():
    return os.getpid()

def get_rsync_process_id(rsync_command, rsync_ppid_callback):
    try:
        # rsyncコマンドを非同期で実行
        rsync_process = subprocess.Popen(rsync_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, bufsize=1)

        # rsyncプロセスのプロセスIDを取得
        rsync_process_id = rsync_process.pid

        # rsyncプロセスの親プロセスIDを取得
        rsync_ppid = os.getppid()

        rsync_ppid_callback(rsync_process_id, rsync_ppid)

        while True:
            line = rsync_process.stdout.readline()
            if not line:
                break
            print(line.strip())  # rsyncの進捗情報を表示

        rsync_process.wait()
    except Exception as e:
        print(f"プロセスIDの取得エラー: {e}")

def rsync_process_id_callback(rsync_process_id, rsync_ppid):
    global rsync_id, rsync_parent_id
    rsync_id = rsync_process_id
    rsync_parent_id = rsync_ppid

def count_colors(file_path, selected_row):
    # 赤と黄色のパターンを示すPatternFillオブジェクトを作成
    red_pattern = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    yellow_pattern = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    # カウンターを初期化
    red_count = 0
    yellow_count = 0

    # エクセルファイルを開く
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    # 選択された行の内容を取得し、セルの色をチェック
    for row in worksheet.iter_rows(min_row=selected_row, max_row=selected_row):
        for cell in row:
            if cell.fill:
                if cell.fill.start_color.rgb == yellow_pattern.start_color.rgb:
                    yellow_count += 1
                elif cell.fill.start_color.rgb == red_pattern.start_color.rgb:
                    red_count += 1

    return red_count, yellow_count

# フォルダのパスを指定
folder_path = '管理ファイルを置く場所の絶対パス'  # フォルダのパスを適切に指定してください

# フォルダ内の.xlsxファイルを検索
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# ファイル一覧を表示し、ユーザーに選択させる
print("ファイルを選択してください:")
for i, excel_file in enumerate(excel_files, start=1):
    print(f"{i}. {excel_file}")

# クライアントソケットを作成
client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

# サーバーのホストとポートを指定
server_host = '送り先（receiverを設置する場所）のIPアドレス'  # サーバーのIPアドレスを指定
server_port = 1234  # サーバーのポート番号を指定　recieverと揃える

try:
    client_socket.connect((server_host, server_port))

    file_choice = int(input("ファイルの選択肢を入力してください: "))
    if 1 <= file_choice <= len(excel_files):
        selected_file = excel_files[file_choice - 1]
        file_path = os.path.join(folder_path, selected_file)
        print(f"選択されたファイル: {selected_file}")

        # B2からB6の選択肢を表示
        selected_row_options = ['カット', 'テロップ', 'エフェクト', 'CG挿入', '音入れ']

        print("B2からB6の行から選択してください:")
        for i, option in enumerate(selected_row_options, start=1):
            print(f"{i}. {option}")

        selected_row_choice = int(input("選択肢を入力してください: "))
        if 1 <= selected_row_choice <= 5:
            selected_row_name = selected_row_options[selected_row_choice - 1]
            selected_row = selected_row_choice + 1  # 選択された行を設定
            print(f"選択された行の内容: {selected_row_name}")

            red_count, yellow_count = count_colors(file_path, selected_row)

            print(f"選択された行（{selected_row_name}）の赤色のセルの総数: {red_count}")
            print(f"選択された行（{selected_row_name}）の黄色のセルの総数: {yellow_count}")

            work_time = 8

            ans = round((((work_time * red_count) / (work_time * yellow_count)) / (work_time * 14)) * 10000)

            print(f"優先度の値：{ans}")

            # rsyncコマンドを実行
            send_folder_path = '送りたいファイルが存在する場所の絶対パス'
            send_files = [f for f in os.listdir(send_folder_path)]

            print("フォルダ内のファイル一覧:")
            for i, file in enumerate(send_files, start=1):
                print(f"{i}. {file}")

            selected_file_choice = int(input("送信したいファイルを選択してください: "))
            if 1 <= selected_file_choice <= len(send_files):
                selected_send_file = send_files[selected_file_choice - 1]
                selected_send_file_path = os.path.join(send_folder_path, selected_send_file)

                remote_host = "送り先のユーザ名@IPアドレス"  # rsync先のリモートホストを指定
                remote_path = "送り先の届いたファイルを配置したい場所の絶対パス"  # rsync先のリモートパスを指定
                rsync_command = f"rsync -av --progress {selected_send_file_path} {remote_host}:{remote_path}/"

                rsync_id = None
                rsync_parent_id = None

                # rsyncのプロセスIDと親プロセスIDを非同期で取得
                rsync_process_thread = threading.Thread(target=get_rsync_process_id, args=(rsync_command, rsync_process_id_callback))
                rsync_process_thread.start()

                # ansの結果とrsyncの親プロセスIDをサーバーに送信
                data_to_send = {
                    'project': selected_file,
                    'task': selected_row_name,
                    'duration': ans
                }

                # rsyncのプロセスIDを取得できるまで待つ
                while rsync_id is None:
                    pass

                data_to_send['select_file'] = selected_send_file  # 追加した部分
                data_to_send['rsync_parent_id'] = rsync_parent_id  # rsyncプロセスIDをデータに追加
                data_to_send['rsync_process_id'] = rsync_id 

                # ansの結果をサーバーに送信
                client_socket.send(str(data_to_send).encode())
                print(f"親：{rsync_parent_id}")
                print(f"子：{rsync_id}")
                print(f"データをサーバーに送信しました.")

            else:
                print("無効な選択です。正しい選択肢を入力してください.")
        else:
            print("無効な選択です。正しい選択肢を入力してください.")
    else:
        print("無効な選択です。正しい選択肢を入力してください.")
except ValueError:
    print("無効な入力です。整数を入力してください.")
